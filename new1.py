import json
import random
import os
import openpyxl

def insert_jsonl_from_subdirs_with_copy(input_dir, excel_file, sheet_name="Sheet1"):
    """
    Traverse subdirectories, read all JSONL files,
    and insert their records into an Excel sheet.

    Existing columns: Case | Note Date | Note
    Adds new columns: File Name | Example ID
    - JSONL notes inserted at random rows.
    - Case and Note Date copied from row above.
    - File Name = JSONL filename without extension.
    """

    # Collect all records from JSONL files
    all_records = []
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]  # remove .jsonl
                with open(file_path, "r", encoding="utf-8") as f:
                    for line in f:
                        rec = json.loads(line)
                        all_records.append({
                            "file_name": clean_name,
                            "example_id": rec.get("example_id"),
                            "note": rec.get("text", "")
                        })

    if not all_records:
        print("⚠️ No .jsonl files found in the directory or subdirectories.")
        return

    # Open workbook
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]

    # Ensure headers exist
    headers = [cell.value for cell in ws[1]]
    if "File Name" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="File Name")
        headers.append("File Name")
    if "Example ID" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Example ID")
        headers.append("Example ID")

    case_col = headers.index("Case") + 1
    date_col = headers.index("Note Date") + 1
    note_col = headers.index("Note") + 1
    file_col = headers.index("File Name") + 1
    id_col = headers.index("Example ID") + 1

    # Insert JSONL records into random rows
    for rec in all_records:
        rand_row = random.randint(2, ws.max_row + 1)  # leave header intact
        ws.insert_rows(rand_row)

        # Copy Case and Note Date from above row
        ws.cell(row=rand_row, column=case_col, value=ws.cell(row=rand_row - 1, column=case_col).value)
        ws.cell(row=rand_row, column=date_col, value=ws.cell(row=rand_row - 1, column=date_col).value)

        # Insert JSONL content
        ws.cell(row=rand_row, column=note_col, value=rec["note"])
        ws.cell(row=rand_row, column=file_col, value=rec["file_name"])
        ws.cell(row=rand_row, column=id_col, value=rec["example_id"])

    wb.save(excel_file)
    print(f"✅ Inserted {len(all_records)} JSONL records into {excel_file} (sheet: {sheet_name})")

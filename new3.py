import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def insert_jsonl_from_subdirs_with_copy(input_dir, excel_file, sheet_name="Sheet1"):
    """
    Traverse subdirectories, read all JSONL files,
    and insert their records into an Excel sheet.

    If the Excel file doesn't exist, create it with headers:
    Case | Note Date | Note | File Name | Example ID

    - JSONL notes inserted at random rows.
    - Case and Note Date copied from row above (if present).
    - File Name = JSONL filename without extension.
    """

    all_records = []
    logging.info(f"Scanning directory: {input_dir}")

    # Collect all records from JSONL files
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]  # remove extension
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_records.append({
                                "file_name": clean_name,
                                "example_id": rec.get("example_id"),
                                "note": rec.get("text", "")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_records)} records total so far")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_records:
        logging.warning("⚠️ No .jsonl files found in the directory or subdirectories.")
        return

    # Try opening workbook, else create new one
    if os.path.exists(excel_file):
        try:
            wb = openpyxl.load_workbook(excel_file)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        except Exception as e:
            logging.error(f"❌ Could not open Excel file {excel_file}: {e}")
            return
    else:
        logging.info(f"Excel file {excel_file} not found → creating a new one.")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"])

    # Ensure headers exist
    headers = [cell.value for cell in ws[1] if cell.value]
    if "File Name" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="File Name")
        headers.append("File Name")
    if "Example ID" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Example ID")
        headers.append("Example ID")

    case_col = headers.index("Case") + 1 if "Case" in headers else None
    date_col = headers.index("Note Date") + 1 if "Note Date" in headers else None
    note_col = headers.index("Note") + 1
    file_col = headers.index("File Name") + 1
    id_col = headers.index("Example ID") + 1

    # Shuffle records for global randomness
    random.shuffle(all_records)
    logging.info(f"Starting insertion of {len(all_records)} records into Excel.")

    # Insert JSONL records into random rows
    for idx, rec in enumerate(all_records, 1):
        rand_row = random.randint(2, ws.max_row + 1)  # leave header intact
        ws.insert_rows(rand_row)

        # Copy Case and Note Date from above row (if available)
        if case_col:
            ws.cell(row=rand_row, column=case_col, value=ws.cell(row=rand_row - 1, column=case_col).value)
        if date_col:
            ws.cell(row=rand_row, column=date_col, value=ws.cell(row=rand_row - 1, column=date_col).value)

        # Insert JSONL content
        ws.cell(row=rand_row, column=note_col, value=rec["note"])
        ws.cell(row=rand_row, column=file_col, value=rec["file_name"])
        ws.cell(row=rand_row, column=id_col, value=rec["example_id"])

        if idx % 100 == 0:
            logging.info(f"Inserted {idx}/{len(all_records)} records...")

    wb.save(excel_file)
    logging.info(f"✅ Successfully inserted {len(all_records)} JSONL records into {excel_file} (sheet: {sheet_name})")

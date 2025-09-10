import json
import random
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
import copy
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def insert_jsonl_new_sheet(input_dir, excel_file, new_sheet_name="CombinedNotes"):
    """
    Create a new Excel sheet combining existing notes and JSONL notes,
    inserting JSONL records at random positions.
    """

    all_jsonl_records = []
    logging.info(f"Scanning directory: {input_dir}")

    # Load JSONL files
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_jsonl_records.append({
                                "Case": None,
                                "Note Date": None,
                                "Note": rec.get("text", ""),
                                "File Name": clean_name,
                                "Example ID": rec.get("example_id")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_jsonl_records)} records total")
                except Exception as e:
                    logging.error(f"Failed to read {file_path}: {e}")

    if not all_jsonl_records:
        logging.warning("No JSONL files found.")
        return

    # Load existing workbook
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        old_sheet = wb.active
    else:
        logging.error(f"Excel file {excel_file} does not exist.")
        return

    # Read existing data and styles
    existing_data = []
    existing_styles = []

    headers = [cell.value for cell in old_sheet[1]]
    if "Case" not in headers:
        headers.append("Case")
    if "Note Date" not in headers:
        headers.append("Note Date")
    if "Note" not in headers:
        headers.append("Note")
    if "File Name" not in headers:
        headers.append("File Name")
    if "Example ID" not in headers:
        headers.append("Example ID")

    for row in old_sheet.iter_rows(min_row=2, max_row=old_sheet.max_row):
        values = [cell.value for cell in row]
        styles = [copy.copy(cell._style) for cell in row]
        # Pad values if missing
        while len(values) < len(headers):
            values.append(None)
            styles.append(None)
        existing_data.append(values)
        existing_styles.append(styles)

    # Combine existing data and JSONL records
    combined_data = existing_data + [[
        rec.get(h, None) for h in headers
    ] for rec in all_jsonl_records]

    # Shuffle combined rows
    random.shuffle(combined_data)

    # Create new sheet
    if new_sheet_name in wb.sheetnames:
        ws = wb[new_sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=new_sheet_name)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write rows
    for row_idx, row_values in enumerate(combined_data, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save workbook
    try:
        wb.save(excel_file)
        wb.close()
        logging.info(f"✅ Successfully created new sheet '{new_sheet_name}' with existing and new notes randomized.")
    except Exception as e:
        logging.error(f"Failed to save Excel file: {e}")


# Example usage
# insert_jsonl_new_sheet("path/to/jsonl/folders", "output.xlsx")

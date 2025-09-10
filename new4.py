import json
import random
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Fill, Alignment, Protection
import copy
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def insert_jsonl_fast(input_dir, excel_file, sheet_name="Sheet1"):
    """
    Fast and robust insertion of JSONL records into an Excel sheet,
    preserving formatting, suitable for large sheets.
    """

    all_records = []
    logging.info(f"Scanning directory: {input_dir}")

    # Load JSONL files
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_records.append({
                                "file_name": clean_name,
                                "example_id": rec.get("example_id"),
                                "note": rec.get("text", "")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_records)} records total")
                except Exception as e:
                    logging.error(f"Failed to read {file_path}: {e}")

    if not all_records:
        logging.warning("No JSONL files found.")
        return

    # Load or create workbook
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"])

    # Ensure headers
    headers = [cell.value for cell in ws[1]]
    required_headers = ["Case", "Note Date", "Note", "File Name", "Example ID"]
    for idx, header in enumerate(required_headers, start=1):
        if header not in headers:
            ws.cell(row=1, column=idx, value=header)
    headers = [cell.value for cell in ws[1]]  # refresh headers

    # Determine column indices
    case_col = headers.index("Case") + 1
    date_col = headers.index("Note Date") + 1
    note_col = headers.index("Note") + 1
    file_col = headers.index("File Name") + 1
    id_col = headers.index("Example ID") + 1

    # Read existing rows and their styles
    data_rows = []
    style_rows = []

    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        styles = [copy.copy(cell._style) for cell in row]  # preserve full cell style
        data_rows.append(values)
        style_rows.append(styles)

    # Shuffle JSONL records
    random.shuffle(all_records)

    # Insert JSONL records randomly in memory
    for rec in all_records:
        insert_idx = random.randint(0, len(data_rows))  # 0 = top, len = bottom
        # Copy Case and Note Date from previous row if exists
        prev_row = data_rows[insert_idx - 1] if insert_idx > 0 else [None]*len(headers)
        new_row_values = [
            prev_row[case_col - 1],
            prev_row[date_col - 1],
            rec["note"],
            rec["file_name"],
            rec["example_id"]
        ]
        data_rows.insert(insert_idx, new_row_values)
        # Copy styles from previous row if exists
        prev_styles = style_rows[insert_idx - 1] if insert_idx > 0 else [None]*len(headers)
        style_rows.insert(insert_idx, prev_styles)

    # Clear existing worksheet
    ws.delete_rows(2, ws.max_row)

    # Write all rows back with styles
    for row_idx, (row_values, row_styles) in enumerate(zip(data_rows, style_rows), start=2):
        for col_idx, (value, style) in enumerate(zip(row_values, row_styles), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if style:
                cell._style = copy.copy(style)  # preserve formatting

    # Save workbook
    try:
        wb.save(excel_file)
        wb.close()
        logging.info(f"✅ Successfully inserted {len(all_records)} JSONL records with formatting preserved.")
    except Exception as e:
        logging.error(f"❌ Failed to save Excel file: {e}")


# Example usage:
# insert_jsonl_fast("path/to/jsonl/folders", "output.xlsx")

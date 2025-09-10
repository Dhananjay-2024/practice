import json
import os
import random
import openpyxl
from openpyxl.styles import PatternFill
import copy
import logging
from datetime import datetime, timedelta

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def insert_jsonl_into_same_sheet(input_dir, excel_file, reference_date_str):
    """
    Inserts JSONL notes into the same Excel sheet, prior to 45 days from reference date.
    - Copies Case and Note Date from row above.
    - Only Note column is highlighted.
    - Other formatting preserved.
    - Header row protected.
    """
    reference_date = datetime.strptime(reference_date_str, "%Y-%m-%d").date()
    threshold_date = reference_date - timedelta(days=45)

    # Load workbook
    if not os.path.exists(excel_file):
        logging.error(f"Excel file {excel_file} does not exist.")
        return

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    required_headers = ["Case", "Note Date", "Note", "File Name", "Example ID"]
    for header in required_headers:
        if header not in headers:
            headers.append(header)

    col_idx = {h: headers.index(h)+1 for h in headers}

    # Load JSONL files
    all_records = []
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                clean_name = os.path.splitext(file_name)[0]
                file_path = os.path.join(root, file_name)
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_records.append({
                                "Note": rec.get("text", ""),
                                "File Name": clean_name,
                                "Example ID": rec.get("example_id")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_records)} total records")
                except Exception as e:
                    logging.error(f"Failed to read {file_path}: {e}")

    if not all_records:
        logging.warning("No JSONL files found.")
        return

    # Determine eligible rows based on Note Date
    eligible_rows = []
    for row in range(2, ws.max_row+1):
        cell_value = ws.cell(row=row, column=col_idx["Note Date"]).value
        if cell_value:
            if isinstance(cell_value, datetime):
                date_val = cell_value.date()
            else:
                try:
                    date_val = datetime.strptime(str(cell_value), "%m-%d-%y").date()
                except:
                    continue
            if date_val <= threshold_date:
                eligible_rows.append(row)

    if not eligible_rows:
        logging.warning("No eligible rows found prior to 45 days from reference date.")
        return

    logging.info(f"{len(eligible_rows)} eligible rows for insertion found.")

    # Highlight for Note column
    highlight_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    random.shuffle(all_records)

    for rec in all_records:
        # Pick random eligible row
        insert_row = random.choice(eligible_rows)
        ws.insert_rows(insert_row)

        # Copy Case and Note Date from row above
        ws.cell(row=insert_row, column=col_idx["Case"], value=ws.cell(row=insert_row-1, column=col_idx["Case"]).value)
        ws.cell(row=insert_row, column=col_idx["Note Date"], value=ws.cell(row=insert_row-1, column=col_idx["Note Date"]).value)

        # Insert JSONL data
        ws.cell(row=insert_row, column=col_idx["Note"], value=rec["Note"])
        ws.cell(row=insert_row, column=col_idx["File Name"], value=rec["File Name"])
        ws.cell(row=insert_row, column=col_idx["Example ID"], value=rec["Example ID"])

        # Highlight Note column
        ws.cell(row=insert_row, column=col_idx["Note"]).fill = highlight_fill

        # Update eligible_rows indices to account for the inserted row
        eligible_rows = [r+1 if r >= insert_row else r for r in eligible_rows]

    wb.save(excel_file)
    wb.close()
    logging.info(f"✅ Successfully inserted {len(all_records)} JSONL notes into existing sheet.")

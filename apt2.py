import os
import json
import random
import logging
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# ---------------- Logging ---------------- #
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler()]
)

# ---------------- Config ---------------- #
DATA_DIR = "data_sub"
EXCEL_FILE = "case_data.xlsx"
NOTE_SHEET = "Note Activity"
ACCOUNT_SHEET = "Account Activity"

SAMPLE_SIZE = 5
OUTPUT_DIR = "case_variants"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---- Case selection ----
# CASE_SELECTION = "all"   # all cases
# CASE_SELECTION = 15      # single case
# CASE_SELECTION = (4, 10) # range inclusive
CASE_SELECTION = "all"


# ---------------- Helpers ---------------- #
def ensure_columns(ws):
    """Ensure required columns exist in Note Activity sheet (except Note)."""
    headers = [cell.value for cell in ws[1]]
    required_cols = ["example_id", "bias"]
    for col in required_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers)+1).value = col
            headers.append(col)
            logging.info(f"Added missing column: {col}")
    return headers


def filter_cases(all_cases):
    """Filter cases based on CASE_SELECTION config."""
    if CASE_SELECTION == "all":
        return all_cases
    elif isinstance(CASE_SELECTION, int):  # single case
        return [CASE_SELECTION] if CASE_SELECTION in all_cases else []
    elif isinstance(CASE_SELECTION, tuple) and len(CASE_SELECTION) == 2:
        low, high = CASE_SELECTION
        return [c for c in all_cases if low <= c <= high]
    else:
        logging.error("Invalid CASE_SELECTION format.")
        return []


def load_bias_records():
    """Load all records grouped by bias (cache for sampling)."""
    bias_records = {}
    for bias_name in os.listdir(DATA_DIR):
        subdir = os.path.join(DATA_DIR, bias_name)
        if not os.path.isdir(subdir):
            continue

        records = []
        for fname in os.listdir(subdir):
            if not fname.endswith(".jsonl"):
                continue
            fpath = os.path.join(subdir, fname)
            logging.info(f"Reading {fpath}")
            with open(fpath, "r", encoding="utf-8") as f:
                for line in f:
                    try:
                        rec = json.loads(line)
                        note_text = f"{rec.get('context','')} {rec.get('question','')}".strip()
                        records.append({
                            "example_id": rec.get("example_id", ""),
                            "Note": note_text,
                            "bias": bias_name
                        })
                    except Exception as e:
                        logging.warning(f"Failed parsing line in {fname}: {e}")

        if records:
            bias_records[bias_name] = records
            logging.info(f"Loaded {len(records)} records for bias={bias_name}")
        else:
            logging.warning(f"No records found in {subdir}")

    return bias_records


def get_case_block(note_df, case_no):
    """Get subset of Note Activity rows for a specific case."""
    case_block = note_df[note_df["Case"] == case_no].copy()
    case_block["Note Date"] = pd.to_datetime(case_block["Note Date"], errors="coerce")
    case_block = case_block.sort_values("Note Date")
    return case_block


def pick_insertion_date(case_block):
    """Pick a reference date for new note (median of existing Note Dates)."""
    if case_block.empty or case_block["Note Date"].isna().all():
        return datetime.today()
    valid_dates = case_block["Note Date"].dropna().sort_values()
    return valid_dates.iloc[len(valid_dates)//2]  # median


# ---------------- Main Logic ---------------- #
def create_case_variants():
    logging.info("Loading workbook for case list...")
    note_df = pd.read_excel(EXCEL_FILE, sheet_name=NOTE_SHEET)

    all_cases = note_df["Case"].dropna().unique().tolist()
    all_cases = [int(c) for c in all_cases if str(c).isdigit()]
    selected_cases = filter_cases(all_cases)

    logging.info(f"Selected cases: {selected_cases}")

    # Load all bias records
    bias_records = load_bias_records()

    # For each case, generate variants
    for case_no in selected_cases:
        logging.info(f"Processing Case {case_no}")
        case_block = get_case_block(note_df, case_no)
        insert_date = pick_insertion_date(case_block)

        for bias_name, records in bias_records.items():
            if not records:
                continue

            # Sample 5 records for this bias and case
            subset = random.sample(records, min(SAMPLE_SIZE, len(records)))
            logging.info(f"Case {case_no}, Bias {bias_name}: {len(subset)} samples")

            for idx, rec in enumerate(subset, start=1):
                logging.info(f"Creating variant {idx} for Case {case_no}, Bias {bias_name}")

                # Reload workbook fresh for each variant
                wb = load_workbook(EXCEL_FILE)
                ws_notes = wb[NOTE_SHEET]

                # Ensure extra columns exist
                headers = ensure_columns(ws_notes)
                col_map = {h: headers.index(h)+1 for h in headers}

                # Find insertion row (first row after insert_date in this case block)
                insert_at = ws_notes.max_row + 1  # default append
                for row in range(2, ws_notes.max_row+1):
                    if ws_notes.cell(row, col_map["Case"]).value == case_no:
                        try:
                            note_date = pd.to_datetime(ws_notes.cell(row, col_map["Note Date"]).value)
                            if note_date >= insert_date:
                                insert_at = row
                                break
                        except Exception:
                            continue

                # Insert row
                ws_notes.insert_rows(insert_at)
                ws_notes.cell(insert_at, col_map["Case"]).value = case_no
                ws_notes.cell(insert_at, col_map["Note Date"]).value = insert_date.strftime("%Y-%m-%d")
                ws_notes.cell(insert_at, col_map["Note"]).value = rec["Note"]
                ws_notes.cell(insert_at, col_map["example_id"]).value = rec["example_id"]
                ws_notes.cell(insert_at, col_map["bias"]).value = rec["bias"]

                # Save as a new variant file
                out_name = f"Case{case_no}_Bias{bias_name}_Variant{idx}.xlsx"
                out_path = os.path.join(OUTPUT_DIR, out_name)
                wb.save(out_path)
                logging.info(f"Saved {out_path}")


# ---------------- Run ---------------- #
if __name__ == "__main__":
    create_case_variants()

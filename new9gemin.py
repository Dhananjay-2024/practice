import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from datetime import datetime, timedelta
import copy

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler()
    ]
)

def parse_note_date(cell_value):
    """
    Robustly parses a cell value into a date object.
    Handles datetime objects and common string formats (m/d/yyyy, MM-DD-YY, YYYY-MM-DD).
    """
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, (int, float)): # Excel stores dates as numbers
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(cell_value)).date()
        except Exception:
            pass

    s_value = str(cell_value).strip()
    for fmt in ["%m/%d/%Y", "%m-%d-%y", "%Y-%m-%d", "%m/%d/%y"]:
        try:
            return datetime.strptime(s_value, fmt).date()
        except ValueError:
            continue
    logging.debug(f"Could not parse date: '{s_value}' with any known format.")
    return None

def consolidate_excel_jsonl_insertion(
    input_dir: str,
    excel_file: str,
    sheet_name: str,
    reference_date_str: str,
    target_prior_intervals: list[int], # List of days prior, e.g., [35, 40, 45]
    highlight_color: str = "FFFACD" # Light yellow
):
    """
    Consolidated function to insert JSONL notes into an Excel sheet.
    New notes are assigned a 'Note Date' based on the 'target_prior_intervals'
    and inserted to maintain the chronological order of the sheet.

    Args:
        input_dir (str): Directory containing JSONL files (can have subdirectories).
        excel_file (str): Path to the Excel file.
        sheet_name (str): The name of the sheet to work with.
        reference_date_str (str): A date string (e.g., "YYYY-MM-DD") to calculate target dates.
        target_prior_intervals (list[int]): A list of integers representing days prior to the
                                             reference date. E.g., [35, 40, 45].
        highlight_color (str): Hex color code for highlighting new notes.
    """
    logging.info(f"Starting consolidated JSONL insertion process for '{excel_file}' (sheet: '{sheet_name}').")

    # 1. Calculate target dates based on intervals
    try:
        reference_date = datetime.strptime(reference_date_str, "%Y-%m-%d").date()
        logging.info(f"Reference Date: {reference_date}")
        
        sorted_intervals = sorted(list(set(target_prior_intervals)))
        calculated_target_dates = {
            days_prior: reference_date - timedelta(days=days_prior)
            for days_prior in sorted_intervals
        }
        for dp, td in calculated_target_dates.items():
            logging.info(f"  - Target Date for {dp} days prior: {td}")
            
    except ValueError as e:
        logging.error(f"❌ Invalid reference_date_str format. Please use YYYY-MM-DD. Error: {e}")
        return

    # 2. Collect all records from JSONL files
    all_jsonl_records = []
    # (The rest of this section is the same as before, loading JSONL data)
    # ... [Code for loading JSONL files remains the same] ...
    logging.info(f"Scanning directory: {input_dir} for .jsonl files...")
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_jsonl_records.append({
                                "Case": None,
                                "Note Date": None,
                                "Note": rec.get("text", ""),
                                "File Name": clean_name,
                                "Example ID": rec.get("example_id")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_jsonl_records)} total records so far.")
                except json.JSONDecodeError as e:
                    logging.error(f"❌ Failed to parse JSONL line in {file_path}: {e}")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_jsonl_records:
        logging.warning("⚠️ No .jsonl files found or no valid records loaded. Exiting.")
        return

    # 3. Load or create workbook and sheet
    # ... [Code for loading/creating Excel file remains the same] ...
    wb = None
    try:
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logging.info(f"Loaded existing sheet '{sheet_name}' from '{excel_file}'.")
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"])
                logging.info(f"Created new sheet '{sheet_name}' in '{excel_file}'.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"])
            logging.info(f"Created new Excel file '{excel_file}' and sheet '{sheet_name}'.")
    except Exception as e:
        logging.error(f"❌ Error opening/creating Excel file '{excel_file}' or sheet '{sheet_name}': {e}")
        if wb: wb.close()
        return

    # 4. Ensure headers exist and get column indices
    # ... [Code for ensuring headers and getting col_map remains the same] ...
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    required_headers = ["Case", "Note Date", "Note", "File Name", "Example ID"]
    for header in required_headers:
        if header not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=header)
            headers.append(header)
    col_map = {h: headers.index(h) + 1 for h in headers}
    if not all(k in col_map for k in required_headers):
        logging.error(f"❌ Critical headers {required_headers} could not be established. Exiting.")
        if wb: wb.close()
        return

    # 5. Read existing data and styles
    # ... [Code for reading existing data/styles into memory remains the same] ...
    existing_rows_data = []
    existing_rows_styles = []
    for r_idx in range(2, ws.max_row + 1):
        row_values = []
        row_styles = []
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            row_values.append(cell.value)
            cell_style_dict = {
                'font': copy.copy(cell.font) if cell.has_style else None,
                'fill': copy.copy(cell.fill) if cell.has_style else None,
                'border': copy.copy(cell.border) if cell.has_style else None,
                'alignment': copy.copy(cell.alignment) if cell.has_style else None,
            }
            row_styles.append(cell_style_dict)
        existing_rows_data.append(row_values)
        existing_rows_styles.append(row_styles)
    logging.info(f"Read {len(existing_rows_data)} existing data rows from sheet.")

    # 6. Assign Target Dates to JSONL records
    random.shuffle(all_jsonl_records)
    new_rows_to_insert = []
    for idx, rec in enumerate(all_jsonl_records):
        # Assign a target date by cycling through the intervals
        days_prior = sorted_intervals[idx % len(sorted_intervals)]
        assigned_note_date = calculated_target_dates[days_prior]
        rec["Note Date"] = assigned_note_date
        
        # Prepare the row data based on the header order
        new_row_data = [rec.get(h, None) for h in headers]
        new_rows_to_insert.append(new_row_data)
    logging.info(f"Assigned target dates to {len(new_rows_to_insert)} new records.")

    # 7. Combine, Sort, and Process the data
    # Create an annotated list to hold all rows for sorting
    # Structure: (date_object, is_new_row_flag, row_data, row_style_dict)
    combined_annotated_rows = []
    note_date_col_idx_0based = col_map["Note Date"] - 1

    for i, row_data in enumerate(existing_rows_data):
        parsed_date = parse_note_date(row_data[note_date_col_idx_0based])
        combined_annotated_rows.append((parsed_date, False, row_data, existing_rows_styles[i]))
        
    for new_row_data in new_rows_to_insert:
        assigned_date = new_row_data[note_date_col_idx_0based]
        combined_annotated_rows.append((assigned_date, True, new_row_data, None)) # Style is placeholder

    # Sort the combined list chronologically. Put rows with no date at the very end.
    combined_annotated_rows.sort(key=lambda x: x[0] if x[0] is not None else datetime.max.date())
    logging.info("Combined and sorted all existing and new rows chronologically.")

    # 8. Post-Sort Processing: Inherit Case/Style and apply highlights
    final_data = []
    final_styles = []
    highlight_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
    case_col_idx_0based = col_map["Case"] - 1
    note_col_idx_0based = col_map["Note"] - 1

    for i, (date, is_new, data, style) in enumerate(combined_annotated_rows):
        if is_new:
            # This is a new row, it needs to inherit Case and Style from the previous row
            if i > 0:
                prev_row_data = final_data[-1]
                prev_row_style = final_styles[-1]
                
                # Inherit Case number
                data[case_col_idx_0based] = prev_row_data[case_col_idx_0based]
                
                # Inherit style, then apply highlight
                style = copy.deepcopy(prev_row_style)
                style[note_col_idx_0based]['fill'] = highlight_fill
            else:
                # This new row is the very first row in the sheet
                style = [{'font': None, 'fill': None, 'border': None, 'alignment': None}] * len(headers)
                style[note_col_idx_0based]['fill'] = highlight_fill
        
        final_data.append(data)
        final_styles.append(style)
    logging.info("Applied Case numbers and styles to new rows post-sorting.")
    
    # 9. Clear existing worksheet data and write back the final sorted data
    # ... [Code for writing back data to Excel is the same as before] ...
    try:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        logging.info("Cleared existing data rows from worksheet.")

        for r_idx, (row_values, row_style_dicts) in enumerate(zip(final_data, final_styles), start=2):
            for c_idx, (value, style_dict) in enumerate(zip(row_values, row_style_dicts), start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if style_dict:
                    if style_dict['font']: cell.font = style_dict['font']
                    if style_dict['fill']: cell.fill = style_dict['fill']
                    if style_dict['border']: cell.border = style_dict['border']
                    if style_dict['alignment']: cell.alignment = style_dict['alignment']
        logging.info(f"Wrote back {len(final_data)} rows to worksheet.")
    except Exception as e:
        logging.error(f"❌ Error writing data back to worksheet: {e}")
        if wb: wb.close()
        return

    # 10. Save workbook
    # ... [Code for saving workbook is the same as before] ...
    try:
        wb.save(excel_file)
        logging.info(f"✅ Successfully inserted {len(new_rows_to_insert)} JSONL records into '{excel_file}' (sheet: '{sheet_name}').")
    except Exception as e:
        logging.error(f"❌ Failed to save Excel file '{excel_file}': {e}")
    finally:
        if wb:
            wb.close()
            logging.info("Workbook closed.")

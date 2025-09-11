import os
import json
import random
import logging
from datetime import timedelta, datetime
import pandas as pd
from openpyxl import load_workbook, Workbook

# ---------------- Logging ---------------- #
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler()]
)

# ---------------- Config ---------------- #
DATA_DIR = "data_sub"
EXCEL_FILE = "case_data.xlsx"
NOTE_SHEET = "Note Activity"
ACCOUNT_SHEET = "Account Activity"

SAMPLE_SIZE = 5
OUTPUT_DIR = "case_variants"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---- Case selection ----
# CASE_SELECTION = "all"   # all cases
# CASE_SELECTION = 15      # single case
# CASE_SELECTION = (4, 10) # range inclusive
CASE_SELECTION = "all"


# ---------------- Helpers ---------------- #
def ensure_columns(ws):
    """Ensure required columns exist in Note Activity sheet (except Note)."""
    headers = [cell.value for cell in ws[1]]
    required_cols = ["example_id", "bias"]
    for col in required_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers)+1).value = col
            headers.append(col)
            logging.info(f"Added missing column: {col}")
    return headers


def filter_cases(all_cases):
    """Filter cases based on CASE_SELECTION config."""
    if CASE_SELECTION == "all":
        return all_cases
    elif isinstance(CASE_SELECTION, int):  # single case
        return [CASE_SELECTION] if CASE_SELECTION in all_cases else []
    elif isinstance(CASE_SELECTION, tuple) and len(CASE_SELECTION) == 2:
        low, high = CASE_SELECTION
        return [c for c in all_cases if low <= c <= high]
    else:
        logging.error("Invalid CASE_SELECTION format.")
        return []


def load_bias_records():
    """Load all records grouped by bias from the data directory (no subdirectories)."""
    DATA_DIR = "data"  # update to new directory
    bias_records = {}
    for fname in os.listdir(DATA_DIR):
        if not fname.endswith(".jsonl"):
            continue
        bias_name = os.path.splitext(fname)[0]
        fpath = os.path.join(DATA_DIR, fname)
        logging.info(f"Reading {fpath}")
        records = []
        with open(fpath, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    rec = json.loads(line)
                    note_text = f"{rec.get('context','')} {rec.get('question','')}".strip()
                    records.append({
                        "example_id": rec.get("example_id", ""),
                        "Note": note_text,
                        "bias": bias_name
                    })
                except Exception as e:
                    logging.warning(f"Failed parsing line in {fname}: {e}")
        if records:
            bias_records[bias_name] = records
            logging.info(f"Loaded {len(records)} records for bias={bias_name}")
        else:
            logging.warning(f"No records found in {fpath}")
    return bias_records


def get_case_block(note_df, case_no):
    """Get subset of Note Activity rows for a specific case."""
    case_block = note_df[note_df["Case"] == case_no].copy()
    case_block["Note Date"] = pd.to_datetime(case_block["Note Date"], errors="coerce")
    case_block = case_block.sort_values("Note Date")
    return case_block


def pick_insertion_date(case_block, queue_date):
    """Pick median Note Date within 90 days before Queue In Date.
    Fallbacks:
      1. 45 days before Queue In Date
      2. Median of all Note Dates for the case
      3. Today (if all else fails)
    """
    if pd.isna(queue_date):
        logging.warning("Queue In Date is missing, falling back to today.")
        return datetime.today()

    start_date = queue_date - timedelta(days=90)
    valid_dates = case_block[
        (case_block["Note Date"] >= start_date) &
        (case_block["Note Date"] <= queue_date)
    ]["Note Date"].dropna().sort_values()

    if not valid_dates.empty:
        median_date = valid_dates.iloc[len(valid_dates)//2]
        return median_date

    # Fallback 1: 45 days before Queue In Date
    fallback = queue_date - timedelta(days=45)
    logging.info(f"No valid dates in window, fallback: {fallback.date()}")

    # Fallback 2: Median of all Note Dates for the case
    all_dates = case_block["Note Date"].dropna().sort_values()
    if not all_dates.empty:
        median_all = all_dates.iloc[len(all_dates)//2]
        logging.info(f"Using median of all Note Dates for safekeeping: {median_all.date()}")
        return median_all

    # Fallback 3: Today
    logging.warning("No Note Dates available, falling back to today.")
    return datetime.today()


# ---------------- Main Logic ---------------- #
def create_case_variants():
    logging.info("Loading workbook for case list...")
    note_df = pd.read_excel(EXCEL_FILE, sheet_name=NOTE_SHEET)
    acct_df = pd.read_excel(EXCEL_FILE, sheet_name=ACCOUNT_SHEET)
    acct_df["Queue In Date"] = pd.to_datetime(acct_df["Queue In Date"], errors="coerce")

    all_cases = note_df["Case"].dropna().unique().tolist()
    all_cases = [int(c) for c in all_cases if str(c).isdigit()]
    selected_cases = filter_cases(all_cases)

    logging.info(f"Selected cases: {selected_cases}")

    # Load all bias records
    bias_records = load_bias_records()

    # Prepare to collect all variants
    all_variant_rows = []
    headers_written = False
    headers_to_keep = None

    for case_no in selected_cases:
        logging.info(f"Processing Case {case_no}")
        case_block = get_case_block(note_df, case_no)

        # Get Queue In Date
        q_date = acct_df.loc[acct_df["Case"] == case_no, "Queue In Date"]
        q_date = q_date.iloc[0] if not q_date.empty else pd.NaT

        insert_date = pick_insertion_date(case_block, q_date)

        for bias_name, records in bias_records.items():
            if not records:
                continue

            subset = random.sample(records, min(SAMPLE_SIZE, len(records)))
            logging.info(f"Case {case_no}, Bias {bias_name}: {len(subset)} samples")

            for idx, rec in enumerate(subset, start=1):
                logging.info(f"Creating variant {idx} for Case {case_no}, Bias {bias_name}")

                # Reload Note Activity for each variant
                wb = load_workbook(EXCEL_FILE)
                ws_notes = wb[NOTE_SHEET]

                headers = ensure_columns(ws_notes)
                col_map = {h: headers.index(h)+1 for h in headers}

                insert_at = ws_notes.max_row + 1
                for row in range(2, ws_notes.max_row+1):
                    if ws_notes.cell(row, col_map["Case"]).value == case_no:
                        try:
                            note_date = pd.to_datetime(ws_notes.cell(row, col_map["Note Date"]).value)
                            if note_date >= insert_date:
                                insert_at = row
                                break
                        except Exception:
                            continue

                ws_notes.insert_rows(insert_at)
                ws_notes.cell(insert_at, col_map["Case"]).value = case_no
                ws_notes.cell(insert_at, col_map["Note Date"]).value = insert_date.strftime("%Y-%m-%d")
                ws_notes.cell(insert_at, col_map["Note"]).value = rec["Note"]
                ws_notes.cell(insert_at, col_map["example_id"]).value = rec["example_id"]
                ws_notes.cell(insert_at, col_map["bias"]).value = rec["bias"]

                # Prepare headers and indexes to keep (excluding example_id and bias)
                if not headers_written:
                    headers_to_keep = [h for h in headers if h not in ("example_id", "bias")]
                    # Add columns for Case, Bias, Variant
                    combined_headers = ["Case", "Bias", "Variant"] + headers_to_keep
                    all_variant_rows.append(combined_headers)
                    headers_written = True
                    col_indexes_to_keep = [headers.index(h) for h in headers_to_keep]

                # Write data rows (excluding example_id and bias)
                for row in ws_notes.iter_rows(min_row=2, values_only=True):
                    filtered_row = [row[i] for i in col_indexes_to_keep]
                    # Add Case, Bias, Variant columns
                    all_variant_rows.append([case_no, bias_name, idx] + filtered_row)

    # Write all variants to a single Excel sheet
    if len(all_variant_rows) > 1:
        wb_all = Workbook()
        ws_all = wb_all.active
        ws_all.title = "All_Case_Variants"
        for row in all_variant_rows:
            ws_all.append(row)
        out_path = os.path.join(OUTPUT_DIR, "All_Case_Variants.xlsx")
        wb_all.save(out_path)
        logging.info(f"Saved all variants to {out_path}")


# ---------------- Run ---------------- #
if __name__ == "__main__":
    create_case_variants()

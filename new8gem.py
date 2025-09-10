import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import copy

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler()
    ]
)

def parse_note_date(cell_value):
    """
    Robustly parses a cell value into a date object.
    Handles datetime objects and common string formats (m/d/yyyy, MM-DD-YY, YYYY-MM-DD).
    """
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, (int, float)): # Excel stores dates as numbers
        try:
            return datetime.fromtimestamp((cell_value - 25569) * 86400).date() # Convert Excel serial date to datetime
        except:
            pass # Fallback to string parsing if conversion fails

    s_value = str(cell_value).strip()
    for fmt in ["%m/%d/%Y", "%m-%d-%y", "%Y-%m-%d", "%m/%d/%y"]:
        try:
            return datetime.strptime(s_value, fmt).date()
        except ValueError:
            continue
    logging.debug(f"Could not parse date: '{s_value}' with any known format.")
    return None

def consolidate_excel_jsonl_insertion(
    input_dir: str,
    excel_file: str,
    sheet_name: str,
    reference_date_str: str,
    days_prior_threshold: int = 45,
    highlight_color: str = "FFFACD" # Light yellow
):
    """
    Consolidated function to insert JSONL notes into an Excel sheet.

    Args:
        input_dir (str): Directory containing JSONL files (can have subdirectories).
        excel_file (str): Path to the Excel file.
        sheet_name (str): The name of the sheet to work with.
        reference_date_str (str): A date string (e.g., "YYYY-MM-DD") to calculate the threshold.
        days_prior_threshold (int): Number of days prior to reference_date for insertion eligibility.
        highlight_color (str): Hex color code for highlighting new notes.
    """
    logging.info(f"Starting consolidated JSONL insertion process for '{excel_file}' (sheet: '{sheet_name}').")

    # 1. Calculate threshold date
    try:
        reference_date = datetime.strptime(reference_date_str, "%Y-%m-%d").date()
        threshold_date = reference_date - timedelta(days=days_prior_threshold)
        logging.info(f"Reference Date: {reference_date} | Insertion Threshold Date ({days_prior_threshold} days prior): {threshold_date}")
    except ValueError as e:
        logging.error(f"❌ Invalid reference_date_str format. Please use YYYY-MM-DD. Error: {e}")
        return

    # 2. Collect all records from JSONL files
    all_jsonl_records = []
    logging.info(f"Scanning directory: {input_dir} for .jsonl files...")
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_jsonl_records.append({
                                "Case": None, # Will be copied from above
                                "Note Date": None, # Will be copied from above
                                "Note": rec.get("text", ""),
                                "File Name": clean_name,
                                "Example ID": rec.get("example_id")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_jsonl_records)} total records so far.")
                except json.JSONDecodeError as e:
                    logging.error(f"❌ Failed to parse JSONL line in {file_path}: {e}")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_jsonl_records:
        logging.warning("⚠️ No .jsonl files found or no valid records loaded. Exiting.")
        return

    # 3. Load or create workbook and sheet
    wb = None
    try:
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logging.info(f"Loaded existing sheet '{sheet_name}' from '{excel_file}'.")
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"]) # Add default headers
                logging.info(f"Created new sheet '{sheet_name}' in '{excel_file}'.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"]) # Add default headers
            logging.info(f"Created new Excel file '{excel_file}' and sheet '{sheet_name}'.")
    except Exception as e:
        logging.error(f"❌ Error opening/creating Excel file '{excel_file}' or sheet '{sheet_name}': {e}")
        if wb: wb.close()
        return

    # 4. Ensure headers exist and get column indices
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    required_headers = ["Case", "Note Date", "Note", "File Name", "Example ID"]
    for header in required_headers:
        if header not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=header)
            headers.append(header) # Update local headers list
    col_map = {h: headers.index(h) + 1 for h in headers}

    if not all(k in col_map for k in ["Case", "Note Date", "Note", "File Name", "Example ID"]):
        logging.error("❌ Critical headers (Case, Note Date, Note, File Name, Example ID) could not be established. Exiting.")
        wb.close()
        return

    # 5. Read existing data and styles
    existing_rows_data = [] # List of lists for cell values
    existing_rows_styles = [] # List of lists for cell styles

    # Start from row 2 to skip header
    for r_idx in range(2, ws.max_row + 1):
        row_values = []
        row_styles = []
        for c_idx in range(1, len(headers) + 1): # Iterate up to the number of columns we care about
            cell = ws.cell(row=r_idx, column=c_idx)
            row_values.append(cell.value)
            # Copy the style object to prevent modification of original
            row_styles.append(copy.copy(cell._style) if hasattr(cell, '_style') else None)
        existing_rows_data.append(row_values)
        existing_rows_styles.append(row_styles)

    logging.info(f"Read {len(existing_rows_data)} existing data rows from sheet.")

    # 6. Determine eligible insertion indices
    # We want to insert *before* rows that meet the date criteria.
    # The index here refers to the position in `existing_rows_data` (0-indexed).
    eligible_insertion_indices = []
    for idx, row_data in enumerate(existing_rows_data):
        note_date_val = row_data[col_map["Note Date"] - 1] # -1 because col_map is 1-indexed
        parsed_date = parse_note_date(note_date_val)
        if parsed_date and parsed_date <= threshold_date:
            eligible_insertion_indices.append(idx)

    if not eligible_insertion_indices:
        logging.warning(f"⚠️ No eligible rows found with 'Note Date' prior to {threshold_date}. JSONL notes will be appended at the end.")
        # If no eligible rows, append to the very end of existing data
        eligible_insertion_indices = [len(existing_rows_data)]

    logging.info(f"Found {len(eligible_insertion_indices)} eligible insertion points.")

    # 7. Shuffle JSONL records for random distribution
    random.shuffle(all_jsonl_records)

    # 8. Insert JSONL records into combined data/styles in memory
    combined_data = existing_rows_data.copy()
    combined_styles = existing_rows_styles.copy()
    highlight_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    for idx, rec in enumerate(all_jsonl_records):
        # Choose a random eligible index
        insert_pos_in_combined = random.choice(eligible_insertion_indices)

        # Copy Case and Note Date from the row *that will be above* the new insertion
        # This is `insert_pos_in_combined - 1` if inserting before an existing row,
        # or the last row if appending.
        if insert_pos_in_combined > 0:
            prev_row_data = combined_data[insert_pos_in_combined - 1]
            rec["Case"] = prev_row_data[col_map["Case"] - 1]
            rec["Note Date"] = prev_row_data[col_map["Note Date"] - 1]
            inherited_style_row = combined_styles[insert_pos_in_combined - 1]
        else: # Inserting at the very beginning (after header)
            rec["Case"] = None
            rec["Note Date"] = None
            inherited_style_row = [None] * len(headers) # No style to inherit

        # Prepare new row values based on current headers order
        new_row_values = [rec.get(h, None) for h in headers]
        combined_data.insert(insert_pos_in_combined, new_row_values)

        # Prepare new row styles: inherit from above, but highlight 'Note'
        new_row_styles = [copy.copy(s) if s else None for s in inherited_style_row]
        note_col_index_0based = col_map["Note"] - 1
        if note_col_index_0based < len(new_row_styles) and new_row_styles[note_col_index_0based]:
            new_row_styles[note_col_index_0based].fill = highlight_fill
        else: # If no style to inherit, create a new style for the note cell
            new_style = openpyxl.styles.Style()
            new_style.fill = highlight_fill
            if note_col_index_0based >= len(new_row_styles): # Extend if necessary
                new_row_styles.extend([None] * (note_col_index_0based - len(new_row_styles) + 1))
            new_row_styles[note_col_index_0based] = new_style._style

        combined_styles.insert(insert_pos_in_combined, new_row_styles)

        # Update eligible indices to reflect the new row insertion
        eligible_insertion_indices = [i + 1 if i >= insert_pos_in_combined else i for i in eligible_insertion_indices]
        # If the chosen index was the last one (append), add the new last index
        if insert_pos_in_combined == len(combined_data) - 1 and len(eligible_insertion_indices) == 0:
             eligible_insertion_indices.append(len(combined_data))


        if (idx + 1) % 100 == 0:
            logging.info(f"Processed {idx + 1}/{len(all_jsonl_records)} JSONL records in memory.")

    logging.info(f"Finished processing {len(all_jsonl_records)} JSONL records in memory.")

    # 9. Clear existing worksheet data (rows 2 onwards) and write back combined data
    try:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1) # Delete all rows except header
        logging.info("Cleared existing data rows from worksheet.")

        for r_idx, (row_values, row_styles) in enumerate(zip(combined_data, combined_styles), start=2):
            for c_idx, (value, style) in enumerate(zip(row_values, row_styles), start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if style:
                    # Apply the copied style object
                    cell._style = copy.copy(style)
        logging.info(f"Wrote back {len(combined_data)} rows to worksheet.")
    except Exception as e:
        logging.error(f"❌ Error writing data back to worksheet: {e}")
        wb.close()
        return

    # 10. Save workbook
    try:
        wb.save(excel_file)
        logging.info(f"✅ Successfully inserted {len(all_jsonl_records)} JSONL records into '{excel_file}' (sheet: '{sheet_name}').")
    except Exception as e:
        logging.error(f"❌ Failed to save Excel file '{excel_file}': {e}")
    finally:
        if wb:
            wb.close()
            logging.info("Workbook closed.")

import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from datetime import datetime, timedelta
import copy

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler()
    ]
)

def parse_note_date(cell_value):
    """
    Robustly parses a cell value into a date object.
    Handles datetime objects and common string formats (m/d/yyyy, MM-DD-YY, YYYY-MM-DD).
    """
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    # openpyxl typically handles Excel's numerical date format automatically when reading cells.
    # If we get a raw number here, it implies it wasn't parsed by openpyxl,
    # so we'll try a common Excel date to datetime conversion (days since 1899-12-30)
    if isinstance(cell_value, (int, float)):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(cell_value)).date()
        except Exception:
            pass

    s_value = str(cell_value).strip()
    for fmt in ["%m/%d/%Y", "%m-%d-%y", "%Y-%m-%d", "%m/%d/%y"]:
        try:
            return datetime.strptime(s_value, fmt).date()
        except ValueError:
            continue
    logging.debug(f"Could not parse date: '{s_value}' with any known format.")
    return None

def consolidate_excel_jsonl_insertion(
    input_dir: str,
    excel_file: str,
    sheet_name: str,
    reference_date_str: str,
    target_prior_intervals: list[int], # New: List of days prior, e.g., [35, 40, 45]
    highlight_color: str = "FFFACD" # Light yellow
):
    """
    Consolidated function to insert JSONL notes into an Excel sheet.
    Insertion points are chosen based on a "regressive lookup" for the "most closer date"
    from the 'Note Date' column, based on specified 'target_prior_intervals'.
    New notes are never inserted at the very end of the sheet.

    Args:
        input_dir (str): Directory containing JSONL files (can have subdirectories).
        excel_file (str): Path to the Excel file.
        sheet_name (str): The name of the sheet to work with.
        reference_date_str (str): A date string (e.g., "YYYY-MM-DD") to calculate target dates.
        target_prior_intervals (list[int]): A list of integers representing days prior to the
                                             reference date. E.g., [35, 40, 45]. The lookup
                                             prioritizes smaller intervals (closer to reference date).
        highlight_color (str): Hex color code for highlighting new notes.
    """
    logging.info(f"Starting consolidated JSONL insertion process for '{excel_file}' (sheet: '{sheet_name}').")

    # 1. Calculate target dates based on intervals
    try:
        reference_date = datetime.strptime(reference_date_str, "%Y-%m-%d").date()
        logging.info(f"Reference Date: {reference_date}")
        
        # Sort intervals to ensure regressive lookup (prioritize smaller 'days_prior')
        sorted_intervals = sorted(list(set(target_prior_intervals)))
        calculated_target_dates = {
            days_prior: reference_date - timedelta(days=days_prior)
            for days_prior in sorted_intervals
        }
        for dp, td in calculated_target_dates.items():
            logging.info(f"  - Target Date for {dp} days prior: {td}")
            
    except ValueError as e:
        logging.error(f"❌ Invalid reference_date_str format. Please use YYYY-MM-DD. Error: {e}")
        return

    # 2. Collect all records from JSONL files
    all_jsonl_records = []
    logging.info(f"Scanning directory: {input_dir} for .jsonl files...")
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_jsonl_records.append({
                                "Case": None, # Will be copied from above
                                "Note Date": None, # Will be copied from above
                                "Note": rec.get("text", ""),
                                "File Name": clean_name,
                                "Example ID": rec.get("example_id")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_jsonl_records)} total records so far.")
                except json.JSONDecodeError as e:
                    logging.error(f"❌ Failed to parse JSONL line in {file_path}: {e}")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_jsonl_records:
        logging.warning("⚠️ No .jsonl files found or no valid records loaded. Exiting.")
        return

    # 3. Load or create workbook and sheet
    wb = None
    try:
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logging.info(f"Loaded existing sheet '{sheet_name}' from '{excel_file}'.")
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"]) # Add default headers
                logging.info(f"Created new sheet '{sheet_name}' in '{excel_file}'.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Case", "Note Date", "Note", "File Name", "Example ID"]) # Add default headers
            logging.info(f"Created new Excel file '{excel_file}' and sheet '{sheet_name}'.")
    except Exception as e:
        logging.error(f"❌ Error opening/creating Excel file '{excel_file}' or sheet '{sheet_name}': {e}")
        if wb: wb.close()
        return

    # 4. Ensure headers exist and get column indices
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    required_headers = ["Case", "Note Date", "Note", "File Name", "Example ID"]
    for header in required_headers:
        if header not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=header)
            headers.append(header) # Update local headers list
    col_map = {h: headers.index(h) + 1 for h in headers}

    if not all(k in col_map for k in ["Case", "Note Date", "Note", "File Name", "Example ID"]):
        logging.error("❌ Critical headers (Case, Note Date, Note, File Name, Example ID) could not be established. Exiting.")
        wb.close()
        return

    # 5. Read existing data and styles
    existing_rows_data = [] # List of lists for cell values
    existing_rows_styles = [] # List of lists of dictionaries for style components

    # Start from row 2 to skip header
    for r_idx in range(2, ws.max_row + 1):
        row_values = []
        row_styles = []
        for c_idx in range(1, len(headers) + 1): # Iterate up to the number of columns we care about
            cell = ws.cell(row=r_idx, column=c_idx)
            row_values.append(cell.value)
            
            # Store style components as a dictionary for easier copying
            cell_style_dict = {
                'font': copy.copy(cell.font) if cell.font else None,
                'fill': copy.copy(cell.fill) if cell.fill else None,
                'border': copy.copy(cell.border) if cell.border else None,
                'alignment': copy.copy(cell.alignment) if cell.alignment else None,
            }
            row_styles.append(cell_style_dict)
        existing_rows_data.append(row_values)
        existing_rows_styles.append(row_styles)

    logging.info(f"Read {len(existing_rows_data)} existing data rows from sheet.")

    # 6. Score all potential insertion points (original indices in existing_rows_data)
    # A lower score is better. Score = (days_prior_interval_index * weight) + (diff_days from target_date)
    # This prioritizes intervals, then closeness within an interval.
    
    scored_insertion_points = [] # List of (priority_score, original_0_indexed_row_position)
    note_date_col_idx_0based = col_map["Note Date"] - 1

    # Iterate through all rows in existing_rows_data (these are potential spots to insert *before*)
    for original_idx, row_data in enumerate(existing_rows_data):
        row_note_date = parse_note_date(row_data[note_date_col_idx_0based])
        
        best_score_for_this_idx = float('inf') # Tracks the best score for inserting *before* this row

        if row_note_date:
            for interval_idx, days_prior in enumerate(sorted_intervals):
                target_date = calculated_target_dates[days_prior]
                diff_days = (row_note_date - target_date).days

                # We want to insert *before* a row whose date is on or after the target_date,
                # and is as close as possible to it.
                if diff_days >= 0: # row_note_date is on or after target_date
                    # Calculate score: interval priority (smaller interval_idx is better) + closeness to target
                    # Weight interval_idx more to ensure regressive lookup priority
                    score = (interval_idx * 1000) + diff_days
                    if score < best_score_for_this_idx:
                        best_score_for_this_idx = score
                        # Once we find a match for a sorted_interval, no need to check further intervals
                        # for *this* row, as those would have higher interval_idx (worse score).
                        # This ensures the "regressive lookup" within the scoring.
            
        # Add to our list of scored points. Even if no perfect date match, assign high score for a physical slot.
        scored_insertion_points.append((best_score_for_this_idx, original_idx))
    
    # Sort all potential insertion points by their score (lowest score first)
    # This gives us a prioritized list of available slots.
    scored_insertion_points.sort(key=lambda x: x[0])

    if not scored_insertion_points or scored_insertion_points[0][0] == float('inf'):
        logging.warning("⚠️ No suitable date-based insertion points found among existing rows. JSONL notes will NOT be inserted.")
        wb.close()
        return

    logging.info(f"Found {len(scored_insertion_points)} potential insertion points, sorted by priority.")

    # 7. Prepare to insert JSONL records based on priority
    combined_data = existing_rows_data.copy()
    combined_styles = existing_rows_styles.copy()
    highlight_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    inserted_count = 0
    # Map original_idx to current_idx (due to insertions shifting things)
    # This list will track how many new rows have been inserted before each original index
    insertion_offsets = [0] * (len(existing_rows_data) + 1) # +1 for an implicit end-marker

    for idx, rec in enumerate(all_jsonl_records):
        if not scored_insertion_points:
            logging.warning(f"Ran out of prioritized insertion points after {inserted_count} records. Remaining {len(all_jsonl_records) - inserted_count} JSONL notes not inserted.")
            break

        # Get the best available insertion point (lowest score)
        score, original_insert_idx = scored_insertion_points.pop(0)

        # Calculate the actual index in the currently modified combined_data
        # This is original_insert_idx + sum of all insertions that happened *before* this original_idx
        current_insert_idx = original_insert_idx + insertion_offsets[original_insert_idx]
        
        # Log if we are using a very low priority spot (fallback)
        if score == float('inf'):
            logging.warning(f"Inserting record {idx+1} at a fallback position (no date-specific match) before original index {original_insert_idx}.")
        else:
            logging.debug(f"Inserting record {idx+1} (score {score}) before original index {original_insert_idx} (current index {current_insert_idx}).")

        # Copy Case and Note Date from the row that will be immediately above the new insertion
        if current_insert_idx > 0:
            prev_row_data = combined_data[current_insert_idx - 1]
            rec["Case"] = prev_row_data[col_map["Case"] - 1]
            rec["Note Date"] = prev_row_data[col_map["Note_Date"] - 1]
            inherited_style_row_dicts = combined_styles[current_insert_idx - 1]
        else: # Inserting at the very beginning (after header, i.e., index 0 of combined_data)
            rec["Case"] = None
            rec["Note Date"] = None
            inherited_style_row_dicts = [{
                'font': None, 'fill': None, 'border': None, 'alignment': None
            }] * len(headers) # No style to inherit, create empty style dicts

        # Prepare new row values based on current headers order
        new_row_values = [rec.get(h, None) for h in headers]
        combined_data.insert(current_insert_idx, new_row_values)

        # Prepare new row styles: inherit from above, but highlight 'Note'
        new_row_style_dicts = [copy.deepcopy(s_dict) if s_dict else {
            'font': None, 'fill': None, 'border': None, 'alignment': None
        } for s_dict in inherited_style_row_dicts]

        note_col_index_0based = col_map["Note"] - 1
        if note_col_index_0based < len(new_row_style_dicts):
            new_row_style_dicts[note_col_index_0based]['fill'] = highlight_fill
        else:
            logging.warning(f"Note column index {note_col_index_0based} out of bounds for new row styles. Highlighting skipped.")

        combined_styles.insert(current_insert_idx, new_row_style_dicts)
        inserted_count += 1

        # Increment offsets for all original indices *after* the current insertion point
        for i in range(original_insert_idx, len(insertion_offsets)):
            insertion_offsets[i] += 1
        
        if inserted_count % 100 == 0:
            logging.info(f"Processed {inserted_count}/{len(all_jsonl_records)} JSONL records in memory.")

    logging.info(f"Finished processing {inserted_count}/{len(all_jsonl_records)} JSONL records in memory.")

    # 8. Clear existing worksheet data (rows 2 onwards) and write back combined data
    try:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        logging.info("Cleared existing data rows from worksheet.")

        for r_idx, (row_values, row_style_dicts) in enumerate(zip(combined_data, combined_styles), start=2):
            for c_idx, (value, style_dict) in enumerate(zip(row_values, row_style_dicts), start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if style_dict:
                    if style_dict['font']: cell.font = style_dict['font']
                    if style_dict['fill']: cell.fill = style_dict['fill']
                    if style_dict['border']: cell.border = style_dict['border']
                    if style_dict['alignment']: cell.alignment = style_dict['alignment']
        logging.info(f"Wrote back {len(combined_data)} rows to worksheet.")
    except Exception as e:
        logging.error(f"❌ Error writing data back to worksheet: {e}")
        if wb: wb.close()
        return

    # 9. Save workbook
    try:
        wb.save(excel_file)
        logging.info(f"✅ Successfully inserted {inserted_count} JSONL records into '{excel_file}' (sheet: '{sheet_name}').")
    except Exception as e:
        logging.error(f"❌ Failed to save Excel file '{excel_file}': {e}")
    finally:
        if wb:
            wb.close()
            logging.info("Workbook closed.")

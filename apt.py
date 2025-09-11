import os
import json
import random
import logging
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------- Logging ---------------- #
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler()]
)

# ---------------- Config ---------------- #
DATA_DIR = "data_sub"
EXCEL_FILE = "case_data.xlsx"
NOTE_SHEET = "Note Activity"
ACCOUNT_SHEET = "Account Activity"

SAMPLE_SIZE = 5

# ---------------- Helpers ---------------- #
def ensure_columns(ws, required_cols):
    """Ensure required columns exist in Note Activity sheet."""
    headers = [cell.value for cell in ws[1]]
    for col in required_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers)+1).value = col
            headers.append(col)
    return headers

def get_case_from_filename(fname):
    """Extract case number from jsonl filename like case4_xxx.jsonl."""
    for part in fname.replace(".jsonl", "").split("_"):
        if part.lower().startswith("case"):
            return int(part.lower().replace("case", ""))
    return None

# ---------------- Main Logic ---------------- #
def insert_notes():
    # Load workbook
    wb = load_workbook(EXCEL_FILE)
    if NOTE_SHEET not in wb.sheetnames or ACCOUNT_SHEET not in wb.sheetnames:
        logging.error("Excel file must contain 'Note Activity' and 'Account Activity' sheets.")
        return
    ws_notes = wb[NOTE_SHEET]
    ws_acct = wb[ACCOUNT_SHEET]

    # Build case -> Queue In Date lookup
    acct_df = pd.read_excel(EXCEL_FILE, sheet_name=ACCOUNT_SHEET)
    acct_df["Queue In Date"] = pd.to_datetime(acct_df["Queue In Date"], errors="coerce")
    case_queue_dates = dict(zip(acct_df["Case"], acct_df["Queue In Date"]))

    # Ensure extra columns
    headers = ensure_columns(ws_notes, ["example_id", "bias"])
    col_map = {h: headers.index(h)+1 for h in headers}  # header -> col index

    # Iterate over subdirectories (bias)
    for bias_name in os.listdir(DATA_DIR):
        subdir = os.path.join(DATA_DIR, bias_name)
        if not os.path.isdir(subdir):
            continue

        all_records = []
        for fname in os.listdir(subdir):
            if not fname.endswith(".jsonl"):
                continue
            case_no = get_case_from_filename(fname)
            if not case_no:
                continue
            fpath = os.path.join(subdir, fname)
            with open(fpath, "r", encoding="utf-8") as f:
                for line in f:
                    try:
                        rec = json.loads(line)
                        note_text = f"{rec.get('context','')} {rec.get('question','')}".strip()
                        all_records.append({
                            "Case": case_no,
                            "example_id": rec.get("example_id", ""),
                            "Note": note_text,
                            "bias": bias_name
                        })
                    except Exception as e:
                        logging.warning(f"Failed parsing line in {fname}: {e}")

        if not all_records:
            continue

        # Sample 5 records
        subset = random.sample(all_records, min(SAMPLE_SIZE, len(all_records)))

        for rec in subset:
            case_no = rec["Case"]
            q_date = case_queue_dates.get(case_no)
            if pd.isna(q_date):
                logging.warning(f"No Queue In Date for case {case_no}")
                continue
            target_date = q_date - timedelta(days=45)  # ~midpoint of 3 months back

            # Find insertion row (keep sorted by Note Date)
            note_dates = [cell.value for cell in ws_notes["B"][1:]]  # column B = Note Date (skip header)
            insert_at = ws_notes.max_row + 1
            for idx, val in enumerate(note_dates, start=2):
                try:
                    if pd.to_datetime(val) >= target_date:
                        insert_at = idx
                        break
                except Exception:
                    continue

            ws_notes.insert_rows(insert_at)
            ws_notes.cell(insert_at, col_map["Case"]).value = case_no
            ws_notes.cell(insert_at, col_map["Note Date"]).value = target_date.strftime("%Y-%m-%d")
            ws_notes.cell(insert_at, col_map["Note"]).value = rec["Note"]
            ws_notes.cell(insert_at, col_map["example_id"]).value = rec["example_id"]
            ws_notes.cell(insert_at, col_map["bias"]).value = rec["bias"]

            logging.info(f"Inserted note for Case {case_no} (bias={rec['bias']}) at row {insert_at}")

    wb.save(EXCEL_FILE)
    logging.info(f"Workbook updated: {EXCEL_FILE}")

# ---------------- Run ---------------- #
if __name__ == "__main__":
    insert_notes()

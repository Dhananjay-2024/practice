import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook
from datetime import datetime, timedelta

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def parse_note_date(cell_value):
    """Parse Note Date from Excel cell (handles datetime or m/d/yyyy string)."""
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    try:
        return datetime.strptime(str(cell_value).strip(), "%m/%d/%Y").date()
    except Exception:
        return None

def insert_jsonl_into_same_sheet(input_dir, excel_file, sheet_name, reference_date_str):
    """
    Insert JSONL notes into the same sheet as existing notes.
    For each note:
      - Look at rows where Note Date is within 90 days before reference_date.
      - Sort those rows by Note Date.
      - Pick the middle row by date and insert the note above it.
    """

    # Convert reference date string to date object
    reference_date = datetime.strptime(reference_date_str, "%m/%d/%Y").date()
    window_start = reference_date - timedelta(days=90)
    logging.info(f"Reference date: {reference_date} | Window start (90 days prior): {window_start}")

    # Collect all records from JSONL files
    all_records = []
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_records.append({
                                "file_name": clean_name,
                                "example_id": rec.get("example_id"),
                                "note": rec.get("text", "")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_records)} records total so far")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_records:
        logging.warning("⚠️ No .jsonl files found in the directory or subdirectories.")
        return

    # Open workbook
    if not os.path.exists(excel_file):
        logging.error(f"Excel file {excel_file} does not exist.")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            logging.error(f"Sheet {sheet_name} not found in {excel_file}.")
            return
    except Exception as e:
        logging.error(f"❌ Could not open Excel file {excel_file}: {e}")
        return

    # Normalize headers
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.strip().lower()] = idx

    required_cols = ["case", "note date", "note"]
    if not all(k in headers for k in required_cols):
        logging.error(f"❌ Required columns {required_cols} not found in sheet headers. Found: {list(headers.keys())}")
        return

    # Add missing optional headers
    if "file name" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="File Name")
        headers["file name"] = len(headers) + 1
    if "example id" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Example ID")
        headers["example id"] = len(headers) + 1

    # Collect candidate rows in 90-day window
    candidate_rows = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=headers["note date"]).value
        date_val = parse_note_date(cell_value)
        logging.info(f"Row {row} - Note Date cell: {cell_value} | Parsed date: {date_val}")
        if date_val and window_start <= date_val <= reference_date:
            candidate_rows.append((row, date_val))

    if not candidate_rows:
        logging.warning("⚠️ No rows found within 90-day window from reference date. Notes will not be inserted.")
        return

    # Sort by date
    candidate_rows.sort(key=lambda x: x[1])
    logging.info(f"Candidate rows within window: {[f'Row {r}, Date {d}' for r, d in candidate_rows]}")

    # Shuffle new records
    random.shuffle(all_records)

    # Insert each record at approx middle row
    for idx, rec in enumerate(all_records, 1):
        mid_index = len(candidate_rows) // 2
        target_row, target_date = candidate_rows[mid_index]

        logging.info(f"Inserting record {idx} above row {target_row} with Note Date {target_date}")

        ws.insert_rows(target_row)

        # Copy Case & Note Date from row above
        ws.cell(row=target_row, column=headers["case"],
                value=ws.cell(row=target_row - 1, column=headers["case"]).value)
        ws.cell(row=target_row, column=headers["note date"],
                value=ws.cell(row=target_row - 1, column=headers["note date"]).value)

        # Insert new note
        ws.cell(row=target_row, column=headers["note"], value=rec["note"])
        ws.cell(row=target_row, column=headers["file name"], value=rec["file_name"])
        ws.cell(row=target_row, column=headers["example id"], value=rec["example_id"])

        if idx % 50 == 0:
            logging.info(f"Inserted {idx}/{len(all_records)} records...")

    wb.save(excel_file)
    logging.info(f"✅ Successfully inserted {len(all_records)} JSONL records into {excel_file} (sheet: {sheet_name})")

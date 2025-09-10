import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook
from datetime import datetime, timedelta

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def parse_note_date(cell_value):
    """Parse Note Date from Excel cell (handles datetime or m/d/yyyy string)."""
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    try:
        return datetime.strptime(str(cell_value).strip(), "%m/%d/%Y").date()
    except Exception:
        return None

def insert_jsonl_into_same_sheet(input_dir, excel_file, sheet_name, reference_date_str):
    """
    Insert JSONL notes into the same sheet as existing notes.
    Notes are inserted before rows where Note Date is <= 45 days before reference_date.
    Case and Note Date are copied from the row above.
    """

    # Convert reference date string to date object
    reference_date = datetime.strptime(reference_date_str, "%m/%d/%Y").date()
    threshold_date = reference_date - timedelta(days=45)
    logging.info(f"Reference date: {reference_date} | Threshold date (45 days prior): {threshold_date}")

    # Collect all records from JSONL files
    all_records = []
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_records.append({
                                "file_name": clean_name,
                                "example_id": rec.get("example_id"),
                                "note": rec.get("text", "")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_records)} records total so far")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_records:
        logging.warning("⚠️ No .jsonl files found in the directory or subdirectories.")
        return

    # Open workbook
    if not os.path.exists(excel_file):
        logging.error(f"Excel file {excel_file} does not exist.")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            logging.error(f"Sheet {sheet_name} not found in {excel_file}.")
            return
    except Exception as e:
        logging.error(f"❌ Could not open Excel file {excel_file}: {e}")
        return

    # Column index mapping
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    if not all(k in headers for k in ["Case", "Note Date", "Note"]):
        logging.error("❌ Required columns (Case, Note Date, Note) not found in sheet headers.")
        return

    # Add missing headers if needed
    if "File Name" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="File Name")
        headers["File Name"] = len(headers) + 1
    if "Example ID" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Example ID")
        headers["Example ID"] = len(headers) + 1

    # Find eligible rows
    eligible_rows = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=headers["Note Date"]).value
        date_val = parse_note_date(cell_value)
        logging.info(f"Row {row} - Note Date cell: {cell_value} | Parsed date: {date_val}")
        if date_val and date_val <= threshold_date:
            eligible_rows.append(row)

    if not eligible_rows:
        logging.warning("⚠️ No eligible rows found prior to 45 days from reference date. Notes will not be inserted.")
        return
    else:
        logging.info(f"{len(eligible_rows)} eligible rows found: {eligible_rows}")

    # Shuffle new records
    random.shuffle(all_records)

    # Insert new notes above eligible rows
    for idx, rec in enumerate(all_records, 1):
        insert_row = random.choice(eligible_rows)
        ws.insert_rows(insert_row)

        # Copy Case & Note Date from row above
        ws.cell(row=insert_row, column=headers["Case"],
                value=ws.cell(row=insert_row - 1, column=headers["Case"]).value)
        ws.cell(row=insert_row, column=headers["Note Date"],
                value=ws.cell(row=insert_row - 1, column=headers["Note Date"]).value)

        # Insert new note
        ws.cell(row=insert_row, column=headers["Note"], value=rec["note"])
        ws.cell(row=insert_row, column=headers["File Name"], value=rec["file_name"])
        ws.cell(row=insert_row, column=headers["Example ID"], value=rec["example_id"])

        if idx % 50 == 0:
            logging.info(f"Inserted {idx}/{len(all_records)} records...")

    wb.save(excel_file)
    logging.info(f"✅ Successfully inserted {len(all_records)} JSONL records into {excel_file} (sheet: {sheet_name})")

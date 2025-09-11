import os
import json
import random
import logging
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook

# ---------------- Logging ---------------- #
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler()]
)

# ---------------- Config ---------------- #
DATA_DIR = "data_sub"
EXCEL_FILE = "case_data.xlsx"
NOTE_SHEET = "Note Activity"
ACCOUNT_SHEET = "Account Activity"

SAMPLE_SIZE = 5

# ---- Case selection ----
# CASE_SELECTION = "all"   # all cases
# CASE_SELECTION = 15      # single case
# CASE_SELECTION = (4, 10) # range inclusive
CASE_SELECTION = "all"

# ---------------- Helpers ---------------- #
def ensure_columns(ws, required_cols):
    """Ensure required columns exist in Note Activity sheet."""
    headers = [cell.value for cell in ws[1]]
    for col in required_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers)+1).value = col
            headers.append(col)
            logging.info(f"Added missing column: {col}")
    return headers

def filter_cases(all_cases):
    """Filter cases based on CASE_SELECTION config."""
    if CASE_SELECTION == "all":
        return all_cases
    elif isinstance(CASE_SELECTION, int):  # single case
        return [CASE_SELECTION] if CASE_SELECTION in all_cases else []
    elif isinstance(CASE_SELECTION, tuple) and len(CASE_SELECTION) == 2:
        low, high = CASE_SELECTION
        return [c for c in all_cases if low <= c <= high]
    else:
        logging.error("Invalid CASE_SELECTION format.")
        return []

def load_bias_records():
    """Load all records grouped by bias (cache for sampling)."""
    bias_records = {}
    for bias_name in os.listdir(DATA_DIR):
        subdir = os.path.join(DATA_DIR, bias_name)
        if not os.path.isdir(subdir):
            continue

        records = []
        for fname in os.listdir(subdir):
            if not fname.endswith(".jsonl"):
                continue
            fpath = os.path.join(subdir, fname)
            logging.info(f"Reading {fpath}")
            with open(fpath, "r", encoding="utf-8") as f:
                for line in f:
                    try:
                        rec = json.loads(line)
                        note_text = f"{rec.get('context','')} {rec.get('question','')}".strip()
                        records.append({
                            "example_id": rec.get("example_id", ""),
                            "Note": note_text,
                            "bias": bias_name
                        })
                    except Exception as e:
                        logging.warning(f"Failed parsing line in {fname}: {e}")

        if records:
            bias_records[bias_name] = records
            logging.info(f"Loaded {len(records)} records for bias={bias_name}")
        else:
            logging.warning(f"No records found in {subdir}")

    return bias_records

# ---------------- Main Logic ---------------- #
def insert_notes():
    logging.info("Loading workbook...")
    wb = load_workbook(EXCEL_FILE)
    if NOTE_SHEET not in wb.sheetnames or ACCOUNT_SHEET not in wb.sheetnames:
        logging.error("Excel file must contain 'Note Activity' and 'Account Activity' sheets.")
        return
    ws_notes = wb[NOTE_SHEET]

    logging.info("Building case -> Queue In Date mapping...")
    acct_df = pd.read_excel(EXCEL_FILE, sheet_name=ACCOUNT_SHEET)
    acct_df["Queue In Date"] = pd.to_datetime(acct_df["Queue In Date"], errors="coerce")
    case_queue_dates = dict(zip(acct_df["Case"], acct_df["Queue In Date"]))

    # Cases to process from Note Activity sheet
    note_df = pd.read_excel(EXCEL_FILE, sheet_name=NOTE_SHEET)
    all_cases = note_df["Case"].dropna().unique().tolist()
    all_cases = [int(c) for c in all_cases if str(c).isdigit()]

    selected_cases = filter_cases(all_cases)
    logging.info(f"Processing cases: {selected_cases}")

    # Ensure extra columns
    headers = ensure_columns(ws_notes, ["example_id", "bias"])
    col_map = {h: headers.index(h)+1 for h in headers}  # header -> col index

    # Load all bias records once
    bias_records = load_bias_records()

    # Iterate each case from Note Activity
    for case_no in selected_cases:
        q_date = case_queue_dates.get(case_no)
        if pd.isna(q_date):
            logging.warning(f"No Queue In Date for case {case_no}")
            continue

        target_date = q_date - timedelta(days=45)
        logging.info(f"Target Note Date for Case {case_no}: {target_date.date()}")

        # Locate block of rows for this case in Note Activity
        case_rows = [
            (idx, ws_notes.cell(idx, col_map["Note Date"]).value)
            for idx in range(2, ws_notes.max_row+1)
            if ws_notes.cell(idx, col_map["Case"]).value == case_no
        ]
        if not case_rows:
            logging.warning(f"No existing rows for Case {case_no}, appending at end")
            case_rows = [(ws_notes.max_row+1, None)]

        # Insert subset notes for each bias, fresh per case
        for bias_name, records in bias_records.items():
            if not records:
                continue
            subset = random.sample(records, min(SAMPLE_SIZE, len(records)))
            logging.info(f"Inserting {len(subset)} notes for Case {case_no}, Bias={bias_name}")

            for rec in subset:
                # Default insert after last row of this case
                insert_at = case_rows[-1][0] + 1

                # Find right place inside this case block
                for idx, note_date in case_rows:
                    try:
                        if pd.to_datetime(note_date) >= target_date:
                            insert_at = idx
                            break
                    except Exception:
                        continue

                logging.info(f"Inserting note for Case {case_no}, bias={bias_name} at row {insert_at}")

                ws_notes.insert_rows(insert_at)
                ws_notes.cell(insert_at, col_map["Case"]).value = case_no
                ws_notes.cell(insert_at, col_map["Note Date"]).value = target_date.strftime("%Y-%m-%d")
                ws_notes.cell(insert_at, col_map["Note"]).value = rec["Note"]
                ws_notes.cell(insert_at, col_map["example_id"]).value = rec["example_id"]
                ws_notes.cell(insert_at, col_map["bias"]).value = rec["bias"]

                logging.debug(f"Inserted record: {rec}")

    wb.save(EXCEL_FILE)
    logging.info(f"Workbook updated: {EXCEL_FILE}")

# ---------------- Run ---------------- #
if __name__ == "__main__":
    insert_notes()

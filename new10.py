import json
import random
import os
import openpyxl
import logging
from openpyxl import Workbook
from datetime import datetime, timedelta

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def parse_note_date(date_val):
    """Parse Note Date in m/d/yyyy format into datetime object."""
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val
    try:
        return datetime.strptime(str(date_val), "%m/%d/%Y")
    except ValueError:
        try:
            return datetime.strptime(str(date_val), "%m-%d-%y")
        except Exception:
            return None

def insert_jsonl_before_45days(input_dir, excel_file, sheet_name, reference_date_str):
    """
    Insert JSONL records into an Excel sheet:
    - Reads JSONL files from subdirectories
    - Finds nearest row <= (reference_date - 45 days)
    - Inserts new notes above that row
    """
    # Parse reference date
    reference_date = datetime.strptime(reference_date_str, "%m/%d/%Y")
    threshold_date = reference_date - timedelta(days=45)
    logging.info(f"Reference date: {reference_date.date()} | Threshold (45 days before): {threshold_date.date()}")

    # Collect all JSONL records
    all_records = []
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_records.append({
                                "file_name": clean_name,
                                "example_id": rec.get("example_id"),
                                "note": rec.get("text", "")
                            })
                    logging.info(f"Loaded {file_name}, total records: {len(all_records)}")
                except Exception as e:
                    logging.error(f"❌ Failed to read {file_path}: {e}")

    if not all_records:
        logging.warning("⚠️ No JSONL records found.")
        return

    # Open Excel workbook
    if not os.path.exists(excel_file):
        logging.error(f"❌ Excel file {excel_file} not found.")
        return

    wb = openpyxl.load_workbook(excel_file)
    if sheet_name not in wb.sheetnames:
        logging.error(f"❌ Sheet {sheet_name} not found in {excel_file}.")
        return
    ws = wb[sheet_name]

    # Map headers
    headers = {cell.value.lower(): idx+1 for idx, cell in enumerate(ws[1]) if cell.value}
    required = ["case", "note date", "note", "file name", "example id"]
    for req in required:
        if req not in headers:
            logging.error(f"❌ Required column '{req}' not found in headers: {headers}")
            return

    case_col = headers["case"]
    date_col = headers["note date"]
    note_col = headers["note"]
    file_col = headers["file name"]
    id_col = headers["example id"]

    # Find best row to insert before
    best_row = None
    best_date = None
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=date_col).value
        date_val = parse_note_date(cell_value)
        logging.info(f"Row {row}: Raw='{cell_value}' Parsed='{date_val}'")
        if date_val and date_val <= threshold_date:
            if best_date is None or date_val > best_date:
                best_row, best_date = row, date_val

    if not best_row:
        logging.warning("⚠️ No eligible rows found before threshold date. Notes will not be inserted.")
        return
    logging.info(f"Chosen row {best_row} with Note Date {best_date.date()} as insertion point.")

    # Shuffle records for randomness
    random.shuffle(all_records)

    # Insert records
    for idx, rec in enumerate(all_records, 1):
        ws.insert_rows(best_row)

        # Copy Case and Note Date from row above
        ws.cell(row=best_row, column=case_col, value=ws.cell(row=best_row - 1, column=case_col).value)
        ws.cell(row=best_row, column=date_col, value=ws.cell(row=best_row - 1, column=date_col).value)

        # Insert new data
        ws.cell(row=best_row, column=note_col, value=rec["note"])
        ws.cell(row=best_row, column=file_col, value=rec["file_name"])
        ws.cell(row=best_row, column=id_col, value=rec["example_id"])

        if idx % 50 == 0:
            logging.info(f"Inserted {idx}/{len(all_records)} records...")

    wb.save(excel_file)
    logging.info(f"✅ Successfully inserted {len(all_records)} notes into {excel_file}, sheet={sheet_name}")

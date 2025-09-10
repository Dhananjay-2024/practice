import json
import random
import os
import openpyxl

def insert_jsonl_notes_randomly(input_dir, excel_file, sheet_name="Sheet1"):
    """
    Insert JSONL notes into the 'Note' column of an existing Excel sheet 
    at random rows without replacing existing notes.
    
    Existing columns: Case | Note Date | Note
    Adds: File Name | Example ID
    Notes are appended as new rows with empty Case/Note Date.
    """

    # Collect all records from JSONL files
    all_records = []
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                with open(file_path, "r", encoding="utf-8") as f:
                    for line in f:
                        rec = json.loads(line)
                        all_records.append({
                            "file_name": file_name,
                            "example_id": rec.get("example_id"),
                            "note": rec.get("text", "")
                        })

    if not all_records:
        print("⚠️ No .jsonl files found in the directory.")
        return

    # Open workbook
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]

    # Add new headers if not already present
    headers = [cell.value for cell in ws[1]]
    if "File Name" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="File Name")
    if "Example ID" not in headers:
        ws.cell(row=1, column=len(headers) + 2, value="Example ID")

    file_col = headers.index("Note") + 2  # File Name column
    id_col = headers.index("Note") + 3    # Example ID column

    # Insert JSONL notes at random rows
    max_row = ws.max_row
    for rec in all_records:
        rand_row = random.randint(2, max_row + 1)  # random row after header
        ws.insert_rows(rand_row)

        ws.cell(row=rand_row, column=headers.index("Note") + 1, value=rec["note"])
        ws.cell(row=rand_row, column=file_col, value=rec["file_name"])
        ws.cell(row=rand_row, column=id_col, value=rec["example_id"])

    wb.save(excel_file)
    print(f"✅ Inserted {len(all_records)} JSONL notes into {excel_file} (sheet: {sheet_name})")

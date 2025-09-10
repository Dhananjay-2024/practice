import json
import random
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import copy
import logging
from datetime import datetime, timedelta

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def insert_jsonl_safe_insert(input_dir, excel_file, reference_date_str, new_sheet_name="CombinedNotes"):
    """
    Create a new sheet combining existing notes and JSONL notes.
    Inserts new notes at random positions prior to 45 days from reference date.
    - Case and Note Date copied from row above.
    - Only Note column highlighted.
    - Other columns inherit formatting.
    - Ensures notes are never inserted above the header row.
    """
    reference_date = datetime.strptime(reference_date_str, "%Y-%m-%d").date()
    threshold_date = reference_date - timedelta(days=45)

    all_jsonl_records = []
    logging.info(f"Scanning directory: {input_dir}")

    # Load JSONL files
    for root, _, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".jsonl"):
                file_path = os.path.join(root, file_name)
                clean_name = os.path.splitext(file_name)[0]
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        for line in f:
                            rec = json.loads(line)
                            all_jsonl_records.append({
                                "Case": None,
                                "Note Date": None,
                                "Note": rec.get("text", ""),
                                "File Name": clean_name,
                                "Example ID": rec.get("example_id")
                            })
                    logging.info(f"Loaded {file_name} → {len(all_jsonl_records)} records total")
                except Exception as e:
                    logging.error(f"Failed to read {file_path}: {e}")

    if not all_jsonl_records:
        logging.warning("No JSONL files found.")
        return

    # Load existing workbook
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        old_sheet = wb.active
    else:
        logging.error(f"Excel file {excel_file} does not exist.")
        return

    # Prepare headers
    headers = [cell.value for cell in old_sheet[1]]
    required_headers = ["Case", "Note Date", "Note", "File Name", "Example ID"]
    for header in required_headers:
        if header not in headers:
            headers.append(header)

    note_col_idx = headers.index("Note") + 1
    note_date_col_idx = headers.index("Note Date") + 1
    case_col_idx = headers.index("Case") + 1

    # Read existing data and styles
    existing_data = []
    existing_styles = []

    for row in old_sheet.iter_rows(min_row=2, max_row=old_sheet.max_row):
        values = [cell.value for cell in row]
        styles = [copy.copy(cell._style) for cell in row]
        while len(values) < len(headers):
            values.append(None)
            styles.append(None)
        existing_data.append(values)
        existing_styles.append(styles)

    # Determine eligible indices based on Note Date in MM-DD-YY
    eligible_indices = []
    for idx, row in enumerate(existing_data):
        cell_date = row[note_date_col_idx - 1]
        if cell_date:
            if isinstance(cell_date, datetime):
                date_val = cell_date.date()
            else:
                try:
                    date_val = datetime.strptime(str(cell_date), "%m-%d-%y").date()
                except:
                    continue
            if date_val <= threshold_date:
                eligible_indices.append(idx)

    # Ensure minimum eligible index is 1 (row 2 in Excel) to protect header
    eligible_indices = [i for i in eligible_indices if i >= 1]

    logging.info(f"Found {len(eligible_indices)} rows eligible for insertion (not above header)")

    # Define highlight style for Note column
    highlight_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    # Insert JSONL records at random eligible positions
    combined_data = existing_data.copy()
    combined_styles = existing_styles.copy()

    for rec in all_jsonl_records:
        if eligible_indices:
            insert_idx = random.choice(eligible_indices)
        else:
            insert_idx = len(combined_data)

        # Copy Case and Note Date from row above if exists
        if insert_idx > 0:
            rec["Case"] = combined_data[insert_idx - 1][case_col_idx - 1]
            rec["Note Date"] = combined_data[insert_idx - 1][note_date_col_idx - 1]

        # Prepare new row values
        new_row_values = [rec.get(h, None) for h in headers]
        combined_data.insert(insert_idx, new_row_values)

        # Inherit style from row above if exists
        inherited_style = combined_styles[insert_idx - 1] if insert_idx > 0 else [None] * len(headers)
        combined_styles.insert(insert_idx, inherited_style)

        # Update eligible indices
        eligible_indices = [i+1 if i>=insert_idx else i for i in eligible_indices]

    # Create new sheet
    if new_sheet_name in wb.sheetnames:
        ws = wb[new_sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=new_sheet_name)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write all rows with styles and highlight only Note column
    for row_idx, (row_values, row_styles) in enumerate(zip(combined_data, combined_styles), start=2):
        for col_idx, (value, style) in enumerate(zip(row_values, row_styles), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if style:
                cell._style = copy.copy(style)
            if col_idx == note_col_idx and row_values[headers.index("File Name")] and row_values[headers.index("Note")] in [r["Note"] for r in all_jsonl_records]:
                cell.fill = highlight_fill

    # Save workbook
    try:
        wb.save(excel_file)
        wb.close()
        logging.info(f"✅ Successfully created new sheet '{new_sheet_name}' with safe insertion, Case & Note Date copied, and Note highlighted.")
    except Exception as e:
        logging.error(f"❌ Failed to save Excel file: {e}")


# Example usage:
# insert_jsonl_safe_insert("path/to/jsonl/folders", "output.xlsx", "2025-09-10")

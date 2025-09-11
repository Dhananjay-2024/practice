import os
import json
import random
import logging
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook

# ---------------- Logging ---------------- #
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler()]
)

# ---------------- Config ---------------- #
DATA_DIR = "data_sub"
EXCEL_FILE = "case_data.xlsx"
NOTE_SHEET = "Note Activity"
ACCOUNT_SHEET = "Account Activity"

SAMPLE_SIZE = 5

# ---- Case selection ----
# Options:
# CASE_SELECTION = "all"
# CASE_SELECTION = 15
# CASE_SELECTION = (4, 10)  # inclusive range
CASE_SELECTION = "all"

# ---------------- Helpers ---------------- #
def ensure_columns(ws, required_cols):
    """Ensure required columns exist in Note Activity sheet."""
    headers = [cell.value for cell in ws[1]]
    for col in required_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers)+1).value = col
            headers.append(col)
    return headers

def filter_cases(all_cases):
    """Filter cases based on CASE_SELECTION config."""
    if CASE_SELECTION == "all":
        return all_cases
    elif isinstance(CASE_SELECTION, int):  # single case
        return [CASE_SELECTION] if CASE_SELECTION in all_cases else []
    elif isinstance(CASE_SELECTION, tuple) and len(CASE_SELECTION) == 2:
        low, high = CASE_SELECTION
        return [c for c in all_cases if low <= c <= high]
    else:
        logging.error("Invalid CASE_SELECTION format.")
        return []

# ---------------- Main Logic ---------------- #
def insert_notes():
    # Load workbook
    wb = load_workbook(EXCEL_FILE)
    if NOTE_SHEET not in wb.sheetnames or ACCOUNT_SHEET not in wb.sheetnames:
        logging.error("Excel file must contain 'Note Activity' and 'Account Activity' sheets.")
        return
    ws_notes = wb[NOTE_SHEET]

    # Build case -> Queue In Date lookup
    acct_df = pd.read_excel(EXCEL_FILE, sheet_name=ACCOUNT_SHEET)
    acct_df["Queue In Date"] = pd.to_datetime(acct_df["Queue In Date"], errors="coerce")
    case_queue_dates = dict(zip(acct_df["Case"], acct_df["Queue In Date"]))

    # Cases to process from Note Activity sheet
    note_df = pd.read_excel(EXCEL_FILE, sheet_name=NOTE_SHEET)
    all_cases = note_df["Case"].dropna().unique().tolist()
    all_cases = [int(c) for c in all_cases if str(c).isdigit()]

    selected_cases = filter_cases(all_cases)
    logging.info(f"Processing cases: {selected_cases}")

    # Ensure extra columns
    headers = ensure_columns(ws_notes, ["example_id", "bias"])
    col_map = {h: headers.index(h)+1 for h in headers}  # header -> col index

    # Iterate each case from Note Activity
    for case_no in selected_cases:
        q_date = case_queue_dates.get(case_no)
        if pd.isna(q_date):
            logging.warning(f"No Queue In Date for case {case_no}")
            continue

        # Collect candidate records from JSONL files for this case
        all_records = []
        for bias_name in os.listdir(DATA_DIR):
            subdir = os.path.join(DATA_DIR, bias_name)
            if not os.path.isdir(subdir):
                continue
            for fname in os.listdir(subdir):
                if not fname.endswith(".jsonl"):
                    continue
                if f"case{case_no}" not in fname.lower():
                    continue
                fpath = os.path.join(subdir, fname)
                with open(fpath, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            rec = json.loads(line)
                            note_text = f"{rec.get('context','')} {rec.get('question','')}".strip()
                            all_records.append({
                                "Case": case_no,
                                "example_id": rec.get("example_id", ""),
                                "Note": note_text,
                                "bias": bias_name
                            })
                        except Exception as e:
                            logging.warning(f"Failed parsing line in {fname}: {e}")

        if not all_records:
            logging.info(f"No JSONL records found for Case {case_no}")
            continue

        # Sample up to 5 records
        subset = random.sample(all_records, min(SAMPLE_SIZE, len(all_records)))

        # Target date = ~45 days before Queue In Date
        target_date = q_date - timedelta(days=45)

        for rec in subset:
            # Find insertion point
            note_dates = [cell.value for cell in ws_notes["B"][1:]]  # column B = Note Date (skip header)
            insert_at = ws_notes.max_row + 1
            for idx, val in enumerate(note_dates, start=2):
                try:
                    if pd.to_datetime(val) >= target_date:
                        insert_at = idx
                        break
                except Exception:
                    continue

            ws_notes.insert_rows(insert_at)
            ws_notes.cell(insert_at, col_map["Case"]).value = case_no
            ws_notes.cell(insert_at, col_map["Note Date"]).value = target_date.strftime("%Y-%m-%d")
            ws_notes.cell(insert_at, col_map["Note"]).value = rec["Note"]
            ws_notes.cell(insert_at, col_map["example_id"]).value = rec["example_id"]
            ws_notes.cell(insert_at, col_map["bias"]).value = rec["bias"]

            logging.info(f"Inserted note for Case {case_no} (bias={rec['bias']}) at row {insert_at}")

    wb.save(EXCEL_FILE)
    logging.info(f"Workbook updated: {EXCEL_FILE}")

# ---------------- Run ---------------- #
if __name__ == "__main__":
    insert_notes()
